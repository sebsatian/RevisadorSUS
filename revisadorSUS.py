import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def extract_number(value):
    """Extracts the numeric part of a cell containing 'number - text'."""
    match = re.match(r"(\d+)", str(value))
    return int(match.group(1)) if match else None

def adjust_column_widths(file_path):
    """Adjust column widths in an Excel file."""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for col in sheet.columns:
        max_length = 0
        column = col[0].column  # Get the column number
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    workbook.save(file_path)

def calculate_and_update_excel(file_path):
    """Process a single Excel file and add calculated columns."""
    try:
        data = pd.read_excel(file_path, sheet_name=0)
    except Exception as e:
        print(f"Error al procesar '{file_path}': {e}")
        return

    # Drop the timestamp column
    if "Marca temporal" in data.columns:
        data = data.drop(columns=["Marca temporal"])

    # Extract numeric responses from the columns
    numeric_columns = data.iloc[:, 1:]  # Exclude email column
    numeric_data = numeric_columns.applymap(extract_number)

    # Calculate sum of responses for odd and even questions
    odd_sum = numeric_data.iloc[:, ::2].sum(axis=1)  # Odd columns
    even_sum = numeric_data.iloc[:, 1::2].sum(axis=1)  # Even columns

    # Calculate intermediate and final results
    odd_minus_5 = odd_sum - 5
    even_minus_25 = 25 - even_sum
    final_score = (odd_minus_5 + even_minus_25) * 2.5

    # Add results to the dataframe
    data["Suma Impares"] = odd_sum
    data["Impares - 5"] = odd_minus_5
    data["Suma Pares"] = even_sum
    data["Pares - 25"] = even_minus_25
    data["Puntaje Final"] = final_score

    # Calculate the average for the last column
    average_row = ['Promedio'] + [''] * (data.shape[1] - 2) + [final_score.mean()]
    data.loc[len(data)] = average_row

    # Save the updated file
    updated_file_name = file_path.replace(".xlsx", "_resultados.xlsx")
    data.to_excel(updated_file_name, index=False)

    # Adjust column widths
    adjust_column_widths(updated_file_name)
    print(f"Archivo procesado y guardado como '{updated_file_name}'.")

def process_all_excels_in_folder():
    """Process all Excel files in the current folder, excluding those with '_resultados' in their name."""
    current_folder = os.getcwd()
    excel_files = [f for f in os.listdir(current_folder) if f.endswith(".xlsx") and "_resultados" not in f]

    if not excel_files:
        print("No se encontraron archivos Excel en la carpeta para procesar.")
        return

    print(f"Procesando {len(excel_files)} archivos Excel encontrados en '{current_folder}':")
    for excel_file in excel_files:
        print(f"Procesando '{excel_file}'...")
        calculate_and_update_excel(excel_file)

# Main execution
if __name__ == "__main__":
    process_all_excels_in_folder()
